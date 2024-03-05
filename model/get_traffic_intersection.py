import numpy as np
from dataclasses import dataclass
import googlemaps
from typing import List, Dict
import openpyxl
from openpyxl.utils import coordinate_to_tuple

ALLOW_GOOGLE_MAP_API = False

@dataclass
class Traffic:
    road_id: str
    road_name: str
    location: str
    traffic_list: list

    def extract_traffic_data(self):
        morning = np.array(self.traffic_list[:6])
        morning[(morning == None) | (morning == '#NUM!')] = 0 
        afternoon = np.array(self.traffic_list[7:])
        afternoon[(afternoon == None) | (afternoon == '#NUM!')] = 0 
        self.morning_truck = morning[:, 0] @ morning[:,1]
        self.morning_car = morning[:, 0] @ morning[:,2]
        self.morning_scooter = morning[:, 0] @ morning[:,3]

        self.afternoon_truck = afternoon[:, 0] @ afternoon[:,1]
        self.afternoon_car = afternoon[:, 0] @ afternoon[:,2]
        self.afternoon_scooter = afternoon[:, 0] @ afternoon[:,3]

    @property
    def road_query(self):
        location_copy = [loc for loc in self.location if loc != None and loc != "-"]
        two_roads = location_copy[:2]
        if two_roads[0] == two_roads[1]:
            two_roads = location_copy[1:3]
        return ' ' .join(two_roads)

    @property
    def to_dict(self):
        return {
            'road_id':self.road_id,
            'road_name':self.road_name,
            'location':self.location,
            'road_query': self.road_query,
            'morning_truck' : self.morning_truck, 
            'morning_car' : self.morning_car, 
            'morning_scooter' : self.morning_scooter, 
            'afternoon_truck' : self.afternoon_truck, 
            'afternoon_car' : self.afternoon_car, 
            'afternoon_scooter' : self.afternoon_scooter, 
        }

def get_table(worksheet, start_row, start_column, num_rows, num_columns,):
    table = []
    for row in range(start_row, start_row + num_rows):
        current_row = []
        for column in range(start_column, start_column + num_columns):
            cell = worksheet.cell(row=row, column=column)
            current_row.append(cell.value)
        table.append(current_row)
    return table

def get_traffic_table(file_name:str) -> Traffic:
    wb = openpyxl.load_workbook(filename=file_name, read_only=True)
    sheet = wb['基本']
    road_id = sheet["B1"].value
    road_name = sheet["B2"].value
    intersection: list[str] = [[cell.value for cell in row] for row in sheet["B14:E14"]][0]
    ## since I am not sure whether the info will match for all files, I search the coordinates of cell
    traffic_table_loc = (13,7)    # default position
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == '站號' and cell.coordinate != 'A1':
                traffic_table_loc = coordinate_to_tuple(cell.coordinate)
    return Traffic(
        road_id         =road_id,
        road_name       =road_name,
        location        =intersection,
        traffic_list    = get_table(sheet, traffic_table_loc[0] + 2 , traffic_table_loc[1] + 4, 12, 4),
    )